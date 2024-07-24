from django.template.loader import render_to_string
from django.shortcuts import render
from django.http import HttpResponse
from django.conf import settings
from django.utils import timezone
from main.models import *

import pandas as pd
import xlwings
import openpyxl
import plotly.express as px
import plotly as plt
import json
import importlib
import os
from datetime import datetime


def return_cell_range(cell_range):  # на вход идет массив объектов-клеток excel, выходит массив значений в этих клетках
    return [cell[0].value for cell in cell_range]


def practice_display(request, practice_id):


    current_practice = practices.objects.get(id=practice_id)
    path = current_practice.practice
    extension = str(path).split(".")[-1]  # расширение файла
    temp_name = str(path)[6:-4]  # имя практики
    tempfile = "media/" + str(request.user) + "_" + temp_name + extension  # собирается имя временного персонального файла



    tables = []


    variables = []
    macro_results = []
    graphs = []


    if extension == "xlsm":  # параметры открытия файля меняются взависимости от формата
        starter_wb = openpyxl.load_workbook(path, read_only=False, keep_vba=True)

    else:
        starter_wb = openpyxl.load_workbook(path)


    input_sheet = starter_wb['Входные данные']

    inputs = []

    input_name = input_sheet['A1']
    number_of_inputs = 1

    while input_name.value is not None:  # Проверяются всех входные данные, а точнее первый столбец листа
        number_of_inputs += 1
        input_name = input_sheet.cell(row=number_of_inputs, column=1)  # Название
        input_type = input_sheet.cell(row=number_of_inputs, column=2).value  # Тип(значение/массив/выбор)
        input_val = input_sheet.cell(row=number_of_inputs, column=3).value  # Значение по умолчанию
        input_measurement = input_sheet.cell(row=number_of_inputs, column=4).value   # Единицы измерения/Варианты выбора
        input_min = input_sheet.cell(row=number_of_inputs, column=5).value  # Минимум
        input_max = input_sheet.cell(row=number_of_inputs, column=6).value  # Максимум
        address = input_sheet.cell(row=number_of_inputs, column=3).coordinate   # Ячейка excel, в которой находится значение переменной
        if input_type == 'выбор':
            input_measurement = input_measurement.split(",")  # Если тип "выбор", создается массив из вариантов выбора
        elif input_type == 'массив':
            input_val = return_cell_range(input_sheet[input_val])  # Если тип "массив", в значение записывается массив значений диапазона
        if input_name.value is not None:  # Если строка не пустая, все данные записываются в массив входных данных
            inputs.append([input_name.value, input_val, input_measurement,
                           input_min, input_max, address, input_type])

    starter_tables_sheet = starter_wb['Начальные таблицы']

    start_tables = []
    starter_names = []

    f = 1
    starter_tables_empty_row_counter = 0
    while starter_tables_empty_row_counter != 1:
        calc_row = starter_tables_sheet.cell(row=f, column=1)
        if starter_tables_sheet.cell(row=f, column=2).value is None:
            starter_names.append(f)  # Собираются строки с именами начальных таблиц
        if calc_row.value is None:
            starter_tables_empty_row_counter += 1
        f += 1


    starter_prev_rows = 1
    for w in range(0, len(starter_names) - 1):  # Для каждого диапазона между строками с именами таблиц
        upper_border = starter_names[w]
        lower_border = starter_names[w + 1]
        name = starter_tables_sheet.cell(row=upper_border, column=1).value

        last_column = 0
        while True:
            if starter_tables_sheet.cell(row=upper_border + 1, column=last_column + 1).value is not None:
                last_column += 1  # Выясняется количество столбцов, которые занимает таблица
            else:
                break

        # Благодаря диапазону, обозначенному двумя строками и крайним правым столбцом, создается датафрейм pandas
        table_dataframe = pd.read_excel('media/' + str(path), sheet_name='Начальные таблицы', skiprows=starter_prev_rows,
                                        nrows=lower_border - upper_border - 2,
                                        usecols=[x for x in range(0, last_column)])

        json_records = table_dataframe.reset_index().to_json(orient='records')
        data2 = json.loads(json_records)
        start_tables.append((name, data2, list(table_dataframe.columns)))
        # Полученный датафрейм вместе с именем добавляется в массив начальных таблиц

        starter_prev_rows = lower_border

    if request.method == 'POST':  # Если пользователь ввел входные данные

        for v in range(0, len(inputs)):  # Проверяются введенные пользователем входные данные
            if inputs[v][6] == 'массив':
                array_range = input_sheet[inputs[v][5]].value
                values_to_be_replaced = list(input_sheet[array_range])
                for q in range(len(values_to_be_replaced)):
                    values_to_be_replaced[q][0].value = request.POST.getlist(inputs[v][0])[q]  # Значения МАССИВА во временном файле меняются на введенные значения
                inputs[v][1] = request.POST.getlist(inputs[v][0])  # Это чтобы на странице введенные значения остались такими же после ввода
            else:
                input_sheet[inputs[v][5]].value = request.POST[inputs[v][0]].replace(".", ",")  # ВЫБОР и ЗНАЧЕНИЕ во временном файле меняются на введенные
                inputs[v][1] = request.POST[inputs[v][0]]   # Это чтобы на странице введенные значения остались такими же после ввода

        
        starter_wb.save(tempfile)  # Временный файл сохраняется под уникальным именем с новыми входными данными

        macros = []
        excel_app = xlwings.App(visible=False)
        excel_book = excel_app.books.open(tempfile)  # с помощью библиотеки xlwings открывается временный файл
        res = excel_book.sheets['Результат']
        macro_counter = 1
        while True:  # Создается массив из имен макросов
            if res[macro_counter, 10].value is not None:
                macros.append(res[macro_counter, 10].value)
                macro_counter += 1
            else:
                break


        for macro in macros:
            current_macro = excel_book.macro(macro)
            current_macro()  # каждым макрос активируется

        excel_book.save()
        excel_book.close()
        excel_app.quit()  # xlwings закрывает книгу и прекращает процесс


        actual_wb = openpyxl.load_workbook(tempfile, data_only=True)  # далее openpyxl открывает временную книгу
        calculation_sheet = actual_wb['Расчеты']
        results = actual_wb['Результат']

        names = []
        tables = []

        macro_results = []
        variables = []
        graph_info = []
        graphs = []

        k = 1
        empty_row_counter = 0  # Отсюда начинается процесс считывания таблиц Расчетов, АБСОЛЮТНО аналогичный процессу считывания Начальных Таблиц
        while empty_row_counter != 1:
            calc_row = calculation_sheet.cell(row=k, column=1)
            if calculation_sheet.cell(row=k, column=2).value is None:
                names.append(k)
            if calc_row.value is None:
                empty_row_counter += 1
            k += 1


        prev_rows = 1
        for m in range(0, len(names) - 1):
            upper_border = names[m]
            lower_border = names[m + 1]
            name = calculation_sheet.cell(row=upper_border, column=1).value

            last_column = 0
            while True:
                if calculation_sheet.cell(row=upper_border + 1, column=last_column + 1).value is not None:
                    last_column += 1

                else:
                    break


            table_dataframe = pd.read_excel(tempfile, sheet_name='Расчеты', skiprows=prev_rows,
                                            nrows=lower_border-upper_border-2,
                                            usecols=[x for x in range(0, last_column)])


            json_records = table_dataframe.reset_index().to_json(orient='records')
            data1 = json.loads(json_records)
            tables.append((name, data1, list(table_dataframe.columns)))
            prev_rows = lower_border
        # Конец считывания таблиц Расчетов


        for d in range(2, macro_counter+1):  # Результаты макросов, активированных ранее, записываются в отдельный массив
            macro_results.append((results.cell(row=d, column=10).value, results.cell(row=d, column=12).value,
                                  results.cell(row=d, column=13).value))


        result_var_counter = 1
        while True:
            if results.cell(row=result_var_counter, column=1).value is not None:
                result_var_counter += 1
            else:
                break


        for n in range(2, result_var_counter):  # Запись результатов типа "значение"
            if results.cell(row=n, column=2).value == "массив":
                variable_range = return_cell_range(results[str(results.cell(row=n, column=3).value)])
                variables.append((results.cell(row=n, column=1).value, variable_range,
                                  results.cell(row=n, column=4).value))
            else:
                variables.append((results.cell(row=n, column=1).value, results.cell(row=n, column=3).value,
                              results.cell(row=n, column=4).value))


        result_graph_counter = 1
        while True:
            if results.cell(row=result_graph_counter, column=8).value is not None:


                result_graph_counter += 1
            else:
                break

        y_axis = []
        for b in range(2, result_graph_counter):  # считывание значений оси Х и нескольких значений оси Y(линий) для графиков
            if results.cell(row=b, column=6).value is not None:  # Если графики еще есть
                name = results.cell(row=b, column=6).value
                x_axis = return_cell_range(calculation_sheet[results.cell(row=b, column=7).value])
                y_axis.append(return_cell_range(calculation_sheet[results.cell(row=b, column=8).value]))
                graph_info.append((name, x_axis, y_axis))  # в массив графиков добавляется имя графика, ось Х и оси Y

                y_axis = []  # массив осей Y обнуляется
            else:
                y_axis.append(return_cell_range(calculation_sheet[results.cell(row=b, column=8).value]))

        for graph in graph_info:
            x_line = graph[1]
            y_graphs = graph[2]

            data = {
                x_line[0]: x_line[1:]
            }
            for each in y_graphs:
                data[each[0]] = each[1:]
            graph_dataframe = pd.DataFrame(data)
            figure = px.line(graph_dataframe, x=x_line[0], y=[each[0] for each in y_graphs],
                             title=graph[0])  # По полученным данным составляются настоящие графики
            actual_graph = plt.offline.plot(figure, auto_open=False, output_type="div")  # График конвертируется в просматриваемый в html-шаблоне вид
            graphs.append(actual_graph)  # Графики собираются в массив
            
        os.remove(tempfile)  # временный файл удаляется с сервера


    context = {'starter_tables': start_tables,
               'inputs': inputs,
               'tables': tables,
               'outputs': variables,
               'graphs': graphs,
               "macros": macro_results}


    current_user = User.objects.get(id=request.user.id)  # Пользователь
    title = practices.objects.select_related('topic_prac').get(id=practice_id).topic_prac.title  # Тема практики

    if request.POST.getlist("testing_mode[]") == ['non_testing']:  # Если практика проводится НЕ в тестовом режиме
        report = render_to_string("practice/practice.html", context=context) + ""
        new_prac_report = PracticeReport(student=current_user, practice=current_practice, report=report,
                                         date=datetime.now())
        new_prac_report.save()  # Сохраняется отчет практики

        new_activity = Activity(user=current_user, datetime=datetime.now(), activity=f'Получил отчет по практике темы "{title}"')
        new_activity.save()  # Факт отправки отчета записывается

    else:  # Если практика проводится в тестовом режиме
        new_activity = Activity(user=current_user, datetime=datetime.now(), activity=f'Посетил практику темы "{title}"')
        new_activity.save()  # Факт посещения записывается



    return render(request, 'practice/practice.html', context)
